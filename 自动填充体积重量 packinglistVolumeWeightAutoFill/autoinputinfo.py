import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

folder_path = os.path.dirname(os.path.abspath(__file__))
pl_keywords = 'pi&inv'
info_keyword = '销售订单'


#save matched path of stored document
print("请确认好已经把装箱单和体积重量信息文件放入此文件夹中")
os.system("pause")

def find_file(folder_path, keywords):
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if all(keyword.lower() in file.lower() for keyword in keywords):
                return os.path.join(root, file)
    return None

# 检查单元格是否属于合并单元格
def is_merged_cell(cell):
    for merged_range in merged_ranges:
        if cell.coordinate in merged_range:  # 检查单元格是否属于某个合并范围
            return True
    return False


pl_file = find_file(folder_path, pl_keywords)
if not pl_file:
    print("未找到pi&inv文件。")
    os.system("pause")
else:
    print(f"找到目标文件：{pl_file}")

    # 查找 info 文件
info_file = find_file(folder_path, [info_keyword])
if not info_file:
    print("未找到体积重量信息 文件。")
    os.system("pause")
else:
    print(f"找到体积重量信息文件文件：{info_file}")

pl_sheets = pd.read_excel(pl_file, sheet_name=None)  # 返回字典形式 {工作表名: DataFrame}


sheet_to_modify = 'PL'
sheet_to_save_image= 'PI'


# 读取 Excel 文件
try:
    wb = load_workbook(pl_file)
    ws = wb[sheet_to_modify]  # 获取需要修改的工作表
    pl_df = pd.read_excel(pl_file, sheet_name=sheet_to_modify)  # 读取 Pandas 表格
    info_df = pd.read_excel(info_file)  # 读取 info 表
except PermissionError:
    print(f"请关闭需要合并的文件:{pl_file} 后重试")
    print(f"请关闭需要合并的文件:{info_file} 后重试")
    os.system("pause")


# 获取合并单元格的范围
merged_ranges = ws.merged_cells.ranges


# 从第 9 行开始（索引为 7）
for i in range(7, len(pl_df)):
    row = pl_df.iloc[i]
    
    # 获取目标文件 C 列的值
    target_value = row[pl_df.columns[2]]  # C 列是第 3 列（索引为 2）
    weight_cell = ws.cell(row=i + 2, column=11)  # K 列（第 11 列），加 2 是因为 Excel 的行索引从 1 开始

    # 检查 K 列是否为合并单元格
    if is_merged_cell(weight_cell):
        print(f"第 {i + 2} 行的 K 列是合并单元格，跳过填充")
        continue

    # 在 info 表中查找对应的行
    matched_row = info_df[info_df['单据行号  (6)'] == target_value]
    
    if not matched_row.empty:
        # 如果匹配成功，提取长、宽、高和毛重
        length = matched_row.iloc[0]['长cm  (80)'] / 100
        width = matched_row.iloc[0]['宽cm  (81)'] / 100
        height = matched_row.iloc[0]['高cm  (82)'] / 100
        weight = matched_row.iloc[0]['毛重/kg  (84)'] 
        
        # 更新原表内容，保持原格式
        if weight_cell.value is None:  # K 列
            weight_cell.value = weight
        if ws.cell(row=i + 2, column=14).value is None:  # N 列（第 14 列）
            ws.cell(row=i + 2, column=14, value=length)
        if ws.cell(row=i + 2, column=15).value is None:  # O 列（第 15 列）
            ws.cell(row=i + 2, column=15, value=width)
        if ws.cell(row=i + 2, column=16).value is None:  # P 列（第 16 列）
            ws.cell(row=i + 2, column=16, value=height)
        print(f"第 {i + 2} 行更新完成")

print("修改完成，即将保存...")

# 保存文件
try:
    wb.save(pl_file)
    print(f"体积重量已保存至文件：{pl_file}")
except PermissionError:
    print(f"无法保存文件。请关闭文件：{pl_file} 后重试。")




os.system("pause")

    
    
