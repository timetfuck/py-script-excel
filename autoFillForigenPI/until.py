import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage  # 用于处理图片
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import get_column_letter
from style import round_thin_border
from openpyxl.utils import range_boundaries


BASE_DIR = os.getcwd()

p2e = pixels_to_EMU

def find_file(folder_path, keyword):
    """
    在指定目录中查找包含关键字的文件。
    
    Args:
        folder_path (str): 要搜索的文件夹路径。
        keyword (str): 文件名中要匹配的关键字。

    Returns:
        str: 找到的文件路径，如果未找到则返回 None。
    """
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if keyword.lower() in file.lower():
                return os.path.join(root, file)
    return None




def xlsx_floating_images(excel_path: str, output_folder: str):
    """提取浮动图片及其锚点位置信息"""
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)  # 创建图片存储目录
    
    # 存储图片信息
    excel_images = {}

    workbook = load_workbook(excel_path, data_only=True)
    for index, sheet in enumerate(workbook.worksheets):
        for img in sheet._images:
            # 生成图片文件名
            img_name = f"{sheet.title}_row{img.anchor._from.row}_col{img.anchor._from.col}.png"
            img_path = os.path.join(output_folder, img_name)

            # 将图片二进制流保存为文件
            PILImage.open(img.ref).convert('RGB').save(img_path)
            print(f"保存图片文件{img_name}到{img_path}")

            # 获取图片锚点位置
            img_info = {
                "initial_row": img.anchor._from.row + 1,  # 起始行
                "initial_col": img.anchor._from.col + 1,  # 起始列
                "ended_row": img.anchor.to.row,  # 结束行
                "ended_col": img.anchor.to.col,  # 结束列
                "image_path": img_path  # 保存的图片路径
            }
            if excel_images.get(index):
                excel_images[index].append(img_info)
            else:
                excel_images[index] = [img_info]
                    


    return excel_images

def get_cell_size(sheet, column, row):
    """
    获取指定单元格的像素宽度和高度
    """
    column_letter = get_column_letter(column) if isinstance(column, int) else column
    col_width = sheet.column_dimensions[column_letter].width or 10  # 默认宽度
    row_height = sheet.row_dimensions[row].height or 15  # 默认高度

    # 转换为像素（粗略估计，1 字符宽 ≈ 7.5 像素，1 行高 ≈ 1.33 倍像素）
    col_pixel_width = int(col_width * 7.5)
    row_pixel_height = int(row_height * 1.33)
    return col_pixel_width, row_pixel_height

def insert_image_in_cell(sheet, image_path, column, row):
    """
    插入图片到指定单元格，计算图片的偏移量，确保居中显示
    """
    # 加载图片
    img = Image(image_path)
    img_width, img_height = img.width, img.height

    # 获取单元格的像素宽高
    col_pixel_width, row_pixel_height = get_cell_size(sheet, column, row)

    # 计算缩放比例，确保图片不会超出单元格范围
    scale = min(col_pixel_width / img_width, row_pixel_height / img_height, 1)
    img.width = int(img_width * scale)
    img.height = int(img_height * scale)

    # 计算偏移量，使图片居中
    col_offset = max(0, (col_pixel_width - img.width) // 2)
    row_offset = max(0, (row_pixel_height - img.height) // 2)

    # 使用 openpyxl 的 Anchor 来精确放置图片
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

    # 设置图片锚点
    img.anchor = TwoCellAnchor(
        _from=AnchorMarker(
            col=ord(column.upper()) - ord("A"),
            colOff=p2e(col_offset),
            row=row - 1,
            rowOff=p2e(row_offset),
        ),
        to=AnchorMarker(
            col=ord(column.upper()) - ord("A"),
            colOff=p2e(col_offset + img.width),
            row=row - 1,
            rowOff=p2e(row_offset + img.height),
        ),
        editAs="oneCell",
    )
    # 添加图片到工作表
    sheet.add_image(img)
  

    print(f"图片 {image_path} 已插入到 {column}{row}，并已居中。")
def find_global_value_name_and_fill(ws, fields_to_fill):
    """
    遍历表格，找到指定字段并在其右边填充值。
    
    Args:
        ws (Worksheet): 工作表对象。
        fields_to_fill (dict): 字典，键为要查找的字段名，值为需要填充的内容。
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in fields_to_fill:
                # 在字段右边单元格填充对应的值
                right_cell = ws.cell(row=cell.row, column=cell.column + 1)
                right_cell.value = fields_to_fill[cell.value]
                print(f"在 {cell.coordinate} 的右边填充值：{fields_to_fill[cell.value]}")

    

def adjust_merged_cells_and_format(ws, start_row):
    """
    插入行后调整合并单元格范围和行列格式。
    """
    # 备份行高
    row_heights = {row: ws.row_dimensions[row].height for row in range(1, ws.max_row + 1) if ws.row_dimensions[row].height}

    # 获取所有合并单元格
    merged_ranges = ws.merged_cells.ranges
    updated_ranges = []

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        # 如果合并单元格在插入行的下方
        if min_row >= start_row:
            # 调整行范围（向下偏移一行）
            min_row += 1
            max_row += 1
        updated_ranges.append((min_col, min_row, max_col, max_row))
    
    # 清除原有的合并单元格
    ws.merged_cells.ranges = []
    # 重新定义调整后的合并单元格
    for new_range in updated_ranges:
        ws.merge_cells(start_row=new_range[1], start_column=new_range[0],
                       end_row=new_range[3], end_column=new_range[2])
    
    # 恢复行高
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height

    # 统一设置新插入行的高度
    ws.row_dimensions[start_row].height = ws.row_dimensions[start_row - 1].height if start_row > 1 else 15

def delete_files_in_directory(directory):
    """
    删除指定目录中的所有文件和子目录。
    """
    # 检查目录是否存在
    if not os.path.exists(directory):
        print(f"目录 {directory} 不存在。")
        return

    # 遍历目录中的所有文件和子目录
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        try:
            # 如果是文件，则删除文件
            if os.path.isfile(file_path):
                os.remove(file_path)
            # 如果是目录，则递归删除目录
            elif os.path.isdir(file_path):
                delete_files_in_directory(file_path)
                os.rmdir(file_path)
        except Exception as e:
            print(f"无法删除 {file_path}: {e}")