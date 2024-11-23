import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import zipfile
import shutil
from xml.etree import ElementTree as ET


import os
folder_path = r'C:\Users\Administrator\Desktop\autoInputFile'
pl_keywords = 'pi&inv'
info_keyword = '销售订单'
output_folder = r'C:\Users\Administrator\Desktop\extracted_images'

#save matched path of stored document
print("请确认好已经把装箱单和体积重量信息文件放入此文件夹中")
os.system("pause")

def find_file(folder_path, keywords):
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if all(keyword.lower() in file.lower() for keyword in keywords):
                return os.path.join(root, file)
    return None

# 检查单元格是否属于合并单元格
def is_merged_cell(cell):
    for merged_range in merged_ranges:
        if cell.coordinate in merged_range:  # 检查单元格是否属于某个合并范围
            return True
    return False


pl_file = find_file(folder_path, pl_keywords)
if not pl_file:
    print("未找到pi&inv文件。")
    os.system("pause")
else:
    print(f"找到目标文件：{pl_file}")

    # 查找 info 文件
info_file = find_file(folder_path, [info_keyword])
if not info_file:
    print("未找到体积重量信息 文件。")
    os.system("pause")
else:
    print(f"找到体积重量信息文件文件：{info_file}")


pl_sheets = pd.read_excel(pl_file, sheet_name=None)  # 返回字典形式 {工作表名: DataFrame}


sheet_to_modify = 'PL'
sheet_to_save_image= 'PI'


# 读取 Excel 文件
try:
    wb = load_workbook(pl_file)
    ws = wb[sheet_to_modify]  # 获取需要修改的工作表
    wi = wb[sheet_to_save_image] #获取到又图片的工作表
    pl_df = pd.read_excel(pl_file, sheet_name=sheet_to_modify)  # 读取 Pandas 表格
    info_df = pd.read_excel(info_file)  # 读取 info 表
except PermissionError:
    print(f"请关闭需要合并的文件:{pl_file} 后重试")
    print(f"请关闭需要合并的文件:{info_file} 后重试")
    os.system("pause")




# 图片获取操作 要记录位置 ------------------------------------
# confirm if the output folder exist
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 解压 Excel 文件
with zipfile.ZipFile(pl_file, 'r') as zip_ref:
    zip_ref.extractall(output_folder)
print(f"文件已解压到：{output_folder}")

# 动态获取命名空间
def get_namespaces(xml_file):
    namespaces = {}
    events = ("start", "start-ns")
    for event, elem in ET.iterparse(xml_file, events):
        if event == "start-ns":
            prefix, uri = elem
            namespaces[prefix] = uri
    return namespaces

extracted_folder = r'C:\Users\Administrator\Desktop\extracted_images'


# 文件夹路径
drawing_folder = os.path.join(extracted_folder, 'xl', 'drawings')
rels_folder = os.path.join(drawing_folder, '_rels')
media_folder = os.path.join(extracted_folder, 'xl', 'media')

# 保存图片位置与文件名的映射
image_positions = []

# 遍历绘图文件夹中的 XML 文件
for drawing_file in os.listdir(drawing_folder):
    if drawing_file.endswith('.xml'):
        drawing_path = os.path.join(drawing_folder, drawing_file)
        print(f"正在处理绘图文件：{drawing_path}")

        # 解析 XML
        tree = ET.parse(drawing_path)
        root = tree.getroot()

        # 动态获取命名空间
        namespaces = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        }

        # 加载对应的 .rels 文件
        rels_file = os.path.join(rels_folder, drawing_file.replace('.xml', '.xml.rels'))
        if not os.path.exists(rels_file):
            print(f"未找到对应的 .rels 文件：{rels_file}")
            continue

        # 解析 .rels 文件
        rels_tree = ET.parse(rels_file)
        rels_root = rels_tree.getroot()
        embed_to_image = {}
        for rel in rels_root.findall("Relationship"):
            embed_id = rel.attrib["Id"]
            target = rel.attrib["Target"]
            if target.startswith("../media/"):
                image_file = target.split("/")[-1]
                embed_to_image[embed_id] = image_file

        # 遍历 xdr:twoCellAnchor 节点
        for anchor in root.findall("xdr:twoCellAnchor", namespaces):
            # 获取起始位置
            from_node = anchor.find("xdr:from", namespaces)
            to_node = anchor.find("xdr:to", namespaces)
            if from_node is not None and to_node is not None:
                row = int(from_node.find("xdr:row", namespaces).text) + 1
                col = int(from_node.find("xdr:col", namespaces).text) + 1

                # 查找图片
                blip = anchor.find(".//a:blip", namespaces)
                if blip is not None:
                    embed_id = blip.attrib.get(f"{{{namespaces['r']}}}embed")
                    if embed_id in embed_to_image:
                        image_file = embed_to_image[embed_id]
                        image_positions.append((row, col, image_file))
                        print(f"图片位置解析成功：行={row}, 列={col}, 文件={image_file}")

# 打印解析结果
print("图片与单元格对应关系：")
for position in image_positions:
    print(position)
# # ----------------------------------------------------------------------------------


# 获取合并单元格的范围
merged_ranges = ws.merged_cells.ranges



# 从第 9 行开始（索引为 7）
for i in range(7, len(pl_df)):
    row = pl_df.iloc[i]
    
    # 获取目标文件 C 列的值
    target_value = row[pl_df.columns[2]]  # C 列是第 3 列（索引为 2）
    weight_cell = ws.cell(row=i + 2, column=11)  # K 列（第 11 列），加 2 是因为 Excel 的行索引从 1 开始

    # 检查 K 列是否为合并单元格
    if is_merged_cell(weight_cell):
        print(f"第 {i + 2} 行的 K 列是合并单元格，跳过填充")
        continue

    # 在 info 表中查找对应的行
    matched_row = info_df[info_df['单据行号  (6)'] == target_value]
    
    if not matched_row.empty:
        # 如果匹配成功，提取长、宽、高和毛重
        length = matched_row.iloc[0]['长cm  (80)'] / 100
        width = matched_row.iloc[0]['宽cm  (81)'] / 100
        height = matched_row.iloc[0]['高cm  (82)'] / 100
        weight = matched_row.iloc[0]['毛重/kg  (84)'] 
        
        # 更新原表内容，保持原格式
        if weight_cell.value is None:  # K 列
            weight_cell.value = weight
        if ws.cell(row=i + 2, column=14).value is None:  # N 列（第 14 列）
            ws.cell(row=i + 2, column=14, value=length)
        if ws.cell(row=i + 2, column=15).value is None:  # O 列（第 15 列）
            ws.cell(row=i + 2, column=15, value=width)
        if ws.cell(row=i + 2, column=16).value is None:  # P 列（第 16 列）
            ws.cell(row=i + 2, column=16, value=height)
        print(f"第 {i + 2} 行更新完成")



print("修改完成，即将保存...")

# 保存修改后的文件（暂时没有图片）
temp_file = pl_file.replace('.xlsx', '_temp.xlsx')
wb.save(temp_file)
print(f"修改后的文件已暂存为：{temp_file}")

# 重新插入图片
wb = load_workbook(temp_file)
ws = wb.active
for row, col, image_file in image_positions:
    img_path = os.path.join(media_folder, image_file)
    if os.path.exists(img_path):
        img = Image(img_path)
        cell = ws.cell(row=row, column=col)
        ws.add_image(img, cell.coordinate)  # 按单元格重新插入图片
        print(f"图片 {image_file} 已重新插入到单元格 {cell.coordinate}")



# 清理临时文件
os.remove(temp_file)
shutil.rmtree(output_folder)
print(f"已删除临时文件夹：{output_folder}")

wb.save(pl_file)
print(f"修改后的文件已保存到：{pl_file}")



os.system("pause")

    
    
