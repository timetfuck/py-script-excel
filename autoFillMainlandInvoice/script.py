import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
from until import find_file, xlsx_floating_images, get_cell_size, insert_image_in_cell, find_global_value_name_and_fill, delete_files_in_directory
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from style import round_thin_border


# get work path
folder_path = os.path.dirname(os.path.abspath(__file__))
template_keyword = "国内订单"
info_keyword = '销售订单'
template_path = os.path.join(folder_path, "template")
output_image_folder =  os.path.join(folder_path, "image")

# 自动生成带时间戳的新文件名
timestamp = datetime.now().strftime("%Y.%m.%d")


# 检查 template 文件夹是否存在
if not os.path.exists(template_path):
    print(f"模板文件夹不存在：{template_path}")
else:
    print(f"模板文件夹路径：{template_path}")
#find 国内清单 template
template_file = find_file(template_path, template_keyword)

#find info file
info_file = find_file(folder_path, info_keyword)

#check whether find this both
if not template_file:
    print(f"未找到模板文件（包含关键词 '{template_keyword}'）。")
    exit()
else:
    print(f"找到模板文件：{template_file}")

if not info_file:
    print(f"未找到信息文件（包含关键词 '{info_keyword}'）。")
    exit()
else:
    print(f"找到信息文件：{info_file}")

# 加载信息文件 (info)
info_df = pd.read_excel(info_file)

# 加载模板文件 (template)
wb = load_workbook(template_file)
ws = wb.active  # 默认选择第一个工作表

# **获取信息文件中的目标列**
# 提取全局信息
orderIdentifier= info_df.iloc[0]["单据编号  (1)"]
clientName = info_df.iloc[0]["客户名称  (103)"]
info_df = pd.read_excel(info_file, dtype={"客户  (2)": str})
clientNumber = info_df.iloc[0]["客户  (2)"]
# 确保输出目录存在
output_folder = os.path.join(folder_path, "output")
if not os.path.exists(output_folder):
    os.makedirs(output_folder)
output_file = os.path.join(output_folder, f"{clientNumber} {orderIdentifier} {clientName} 国内清单 {timestamp}.xlsx")

recepient = {"收件人：": clientName}
orderIdentifierDic = {"订单号：": orderIdentifier}
timeDic = {"出货日期：": timestamp}

find_global_value_name_and_fill(ws, recepient)
find_global_value_name_and_fill(ws, orderIdentifierDic)
find_global_value_name_and_fill(ws, timeDic)

images_info = xlsx_floating_images(info_file, output_image_folder)
print(images_info)

# 读取信息文件
def read_material_info(info_file):
    info_df = pd.read_excel(info_file)
    material_data = info_df[[
        "单据行号  (6)", "物料编码  (7)", "物料名称  (8)", "英文名  (11)",
        "系统图片  (12)", "品牌  (14)", "数量  (13)", "单位  (15)",
        "单价  (27)", "不含税金额  (29)", "特殊要求及其他  (16)"
    ]].fillna("")
    print(material_data)
    return material_data

# 动态插入行并写入数据
def fill_template_with_material_info(ws, material_data, output_file):
    gn_number_format ='¥#,##0.00'  
    start_row = 8  # 数据写入的起始行
    current_row = start_row
    for i, row in material_data.iterrows():
        # 动态插入行，确保金额部分不被覆盖
        ws.insert_rows(current_row)
        row_number = row["单据行号  (6)"]
        info_current_row = i+2
        #设定单元格长度宽度
        cell_height = 36
        ws.row_dimensions[current_row].height = cell_height      # 设置第 current 行高度为 36
        # 写入数据
        if row_number: 
            ws[f"A{current_row}"] = row_number
            # set number of order alignent to center
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f"A{current_row}"].border = round_thin_border 

            ws[f"B{current_row}"] = row["物料编码  (7)"][5:]
            ws[f'B{current_row}'].alignment = Alignment(vertical='center')
            ws[f"B{current_row}"].border = round_thin_border 
            #set width of coloum to 30
            ws.column_dimensions['B'].width = 30  #列宽为30
            ws[f"C{current_row}"] = row["物料名称  (8)"].replace("/有图片","")
            ws[f"C{current_row}"].border = round_thin_border 
            ws[f'C{current_row}'].alignment = Alignment(vertical='center')

            ws[f"E{current_row}"] = row["品牌  (14)"]
            ws[f"E{current_row}"].border = round_thin_border 
            ws[f'E{current_row}'].alignment = Alignment(vertical='center')

            ws[f"F{current_row}"] = row["数量  (13)"]
            ws[f'F{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f"F{current_row}"].border = round_thin_border 

            ws[f"G{current_row}"] = row["单位  (15)"]
            ws[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f"G{current_row}"].border = round_thin_border 

            ws[f"H{current_row}"] = row["单价  (27)"]
            ws[f"H{current_row}"].number_format = gn_number_format 
            ws[f'H{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f"H{current_row}"].border = round_thin_border 

            ws[f"I{current_row}"] = row["不含税金额  (29)"]
            ws[f"I{current_row}"].number_format = gn_number_format
            ws[f'I{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f"I{current_row}"].border = round_thin_border 

            ws[f"J{current_row}"] = row["特殊要求及其他  (16)"]
        else:
            ws[f"H{current_row}"] = "金额"  
            ws[f"H{current_row}"].border = round_thin_border 
            ws[f"H{current_row}"].font = Font(size=16, bold=True)  
            ws[f"H{current_row}"].alignment = Alignment(horizontal="center", vertical="center") 

            ws[f"I{current_row}"].value = f"=sum(I{start_row}:I{current_row - 1})" 
            ws[f"I{current_row}"].alignment = Alignment(horizontal="center", vertical="center") 
            ws[f"I{current_row}"].font = Font(size=16, bold=True)  
            ws[f"I{current_row}"].number_format = '¥#,##0.00' 
            ws[f"I{current_row}"].border = round_thin_border 
            
            ws[f"F{current_row}"].value = f"=sum(F{start_row}:F{current_row - 1})" 
            ws[f"F{current_row}"].alignment = Alignment(horizontal="center", vertical="center") 
            ws[f"F{current_row}"].font = Font(size=16, bold=True)  
            ws[f"F{current_row}"].border = round_thin_border 
            
      # 插入图片
        matching_image = {}
        for sheet_index, images in images_info.items():
            for img in images:
                if img["initial_row"] == info_current_row:
                    matching_image = img
        if matching_image:
            img_path = matching_image["image_path"]
            if os.path.exists(img_path):  # 确保图片文件存在
                ws[f"D{current_row}"].border = round_thin_border 
                insert_image_in_cell(ws, img_path,"D", current_row)
        current_row += 1  # 移动到下一行

    print(f"数据已成功写入并保存到 {output_file}")

# 主程序
material_data = read_material_info(info_file)
fill_template_with_material_info(ws, material_data, output_file)

# 保存文件
wb.save(output_file)
delete_files_in_directory(output_image_folder)
os.removedirs(output_image_folder)


